from openpyxl import load_workbook

LAB_NUM=3
NAME='FileForEnteringLab3Marks.xlsx'
wb=load_workbook(NAME, data_only=True)
ws = wb['2017S2CompSci111']


upis=[ws.cell(row=x, column=8) for x in range(3,368)]
fail_list=[]
times={}
while True:
    count=0
    t=input('Name: ')
    if t.lower()=='done':break
    for upi in upis:
        if t in upi.value:
            count+=1
            mark=ws.cell(row=upi.row, column=9+LAB_NUM)
            time=ws.cell(row=upi.row, column=3).internal_value
    if count!=1:
        print('INVALID NAME, APPEARANCE:',count)
        fail_list.append(t)
        continue
    while True:
        try:
            s=int(input('Mark: '))
            if s<0 or s>100: raise
            break
        except:
            print("INVALID MARK")
            
    mark.value=s
    try: times[time].append(t)
    except: times[time]=[t]
    wb.save(NAME)


    
c=0
for time in times:
    c+=len(times[time])
    print(time,', total: ',len(times[time]),sep='')
    
          

    
    
    
            
    


wb.save('test.xlsx')

    
            
            



